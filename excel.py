from openpyxl import load_workbook


workbook = load_workbook(filename="test_workbook.xlsx")

workbook.create_sheet("TestSheet2")
print(workbook.sheetnames)

workbook.create_sheet("TestSheet3", 0)

test_sheet_3 = workbook["TestSheet3"]

workbook.copy_worksheet(test_sheet_3)

print(workbook.sheetnames)
workbook.remove(test_sheet_3)
print(workbook.sheetnames)

workbook.save(filename="test_workbook.xlsx")
