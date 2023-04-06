from openpyxl import load_workbook


source_file ="source_workbook.xlsx"

source_workbook = load_workbook(source_file)
source_worksheet = source_workbook.worksheets[0]

destination_file ="destination_workbook.xlsx"
destination_workbook = load_workbook(destination_file)
destination_worksheet = destination_workbook.active

max_row = source_worksheet.max_row

for row in range (1, max_row +1):
    for column in range (1, 3):

        c =source_worksheet.cell(row = row, column = column)

        destination_worksheet.cell(row = row, column = column).value = c.value
destination_workbook.save(destination_file)

